import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
import io

st.set_page_config(page_title="Absensi OJT Digital", layout="wide", page_icon="📝")

# --- CUSTOM CSS UNTUK TAMPILAN ELEGAN ---
st.markdown("""
    <style>
    /* Styling tombol agar terlihat lebih premium */
    div.stButton > button:first-child {
        border-radius: 8px;
        font-weight: 500;
        transition: all 0.3s ease;
    }
    div.stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    /* Memperhalus tampilan Metric (Statistik) */
    div[data-testid="metric-container"] {
        background-color: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.1);
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    </style>
""", unsafe_allow_html=True)

def get_color_str(fill):
    """Mengekstrak kode warna dari sel Excel."""
    if not fill: return None
    if fill.fgColor:
        if fill.fgColor.type == 'rgb': return fill.fgColor.rgb
        elif fill.fgColor.type == 'indexed': return str(fill.fgColor.indexed)
    if fill.start_color and fill.start_color.index:
        return str(fill.start_color.index)
    return None

# --- INISIALISASI SESSION STATE ---
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False
if 'weeks_data' not in st.session_state:
    st.session_state.weeks_data = {}
if 'students_data' not in st.session_state:
    st.session_state.students_data = {}
if 'absensi_df' not in st.session_state:
    st.session_state.absensi_df = pd.DataFrame()
if 'current_viewed_week' not in st.session_state:
    st.session_state.current_viewed_week = None

# ==========================================
# SIDEBAR KONTROL UTAMA
# ==========================================
with st.sidebar:
    st.title("⚙️ Pengaturan OJT")
    st.markdown("---")
    
    st.markdown("### 1. Upload Jadwal")
    uploaded_file = st.file_uploader("Upload OJT 2026.xlsx", type=["xlsx"])
    
    # Tombol Reset
    if st.session_state.data_loaded:
        if st.button("🔄 Reset Data (Upload Ulang)", use_container_width=True):
            st.session_state.data_loaded = False
            st.session_state.weeks_data = {}
            st.session_state.students_data = {}
            st.rerun()

# Eksekusi Pembacaan File (Hanya dilakukan sekali saat upload)
if uploaded_file is not None and not st.session_state.data_loaded:
    with st.spinner("Memproses Excel..."):
        try:
            wb_source = openpyxl.load_workbook(uploaded_file, data_only=True)
            ws_source = wb_source['JADWAL']
            
            # Ekstrak Jadwal (Logika Ganda untuk minggu pengulangan)
            weeks = {}
            for row in range(4, 60):
                for col in [9, 11]:
                    cell = ws_source.cell(row=row, column=col)
                    if type(cell.value) in [int, float]:
                        c_str = get_color_str(cell.fill)
                        d_val = ws_source.cell(row=row, column=col+1).value
                        weeks[int(cell.value)] = {
                            'color': c_str,
                            'date': str(d_val) if d_val else "-"
                        }
            st.session_state.weeks_data = weeks

            # Ekstrak Data Siswa
            classes = {col: ws_source.cell(row=4, column=col).value for col in range(3, 8)}
            students = {}
            for row in range(5, 60):
                for col in range(3, 8):
                    cell = ws_source.cell(row=row, column=col)
                    val = str(cell.value).strip() if cell.value else ""
                    if len(val) > 2:
                        students[val] = {
                            'kelas': str(classes.get(col, '')).strip(),
                            'color': get_color_str(cell.fill)
                        }
            st.session_state.students_data = students
            st.session_state.data_loaded = True
            st.rerun()
            
        except Exception as e:
            st.sidebar.error(f"Error membaca file: {e}")

# ==========================================
# AREA UTAMA (MAIN AREA)
# ==========================================
if not st.session_state.data_loaded:
    # Tampilan awal jika belum ada file
    st.markdown("<h1 style='text-align: center; color: #4A90E2;'>Sistem Absensi OJT Digital</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #888;'>Silakan unggah file jadwal induk Anda pada menu di sebelah kiri untuk memulai.</p>", unsafe_allow_html=True)
    
else:
    if not st.session_state.weeks_data:
        st.error("Jadwal mingguan tidak terdeteksi di file.")
    else:
        # Pilihan Minggu di Sidebar (Hanya muncul kalau file sudah diupload)
        with st.sidebar:
            st.markdown("### 2. Pilih Minggu OJT")
            week_options = sorted(list(st.session_state.weeks_data.keys()))
            selected_week = st.selectbox("Tampilkan Data Minggu Ke-:", week_options)
            
            target_color = st.session_state.weeks_data[selected_week]['color']
            target_date = st.session_state.weeks_data[selected_week]['date']
            
            st.info(f"📅 **Periode:**\n{target_date}")
            
        # --- TAB LAYOUT (MEMBUAT TAMPILAN LEBIH RAPI) ---
        tab_absensi, tab_kelola = st.tabs(["📝 Isi Absensi", "🛠️ Kelola Pertukaran Jadwal"])
        
        with tab_kelola:
            st.subheader("Pertukaran / Pemindahan Siswa")
            st.write("Gunakan fitur ini jika ada siswa yang pindah atau bertukar jadwal minggu OJT.")
            
            all_student_names = sorted(list(st.session_state.students_data.keys()))
            c_tambah, c_tukar = st.columns(2)
            
            with c_tambah:
                st.markdown("#### Pindah ke Minggu Ini")
                st.caption("Pilih siswa dari minggu lain untuk dimasukkan ke daftar minggu ini.")
                s_tambah = st.selectbox("Siswa yang ingin dipindah:", ["-- Pilih Siswa --"] + all_student_names, key="add_s")
                if st.button("Tambahkan Siswa", type="secondary") and s_tambah != "-- Pilih Siswa --":
                    st.session_state.students_data[s_tambah]['color'] = target_color
                    st.success(f"{s_tambah} berhasil dipindah!")
                    st.rerun()

            with c_tukar:
                st.markdown("#### Tukar Posisi (Swap)")
                st.caption("Tukar jadwal Siswa A (Minggu ini) dengan Siswa B (Minggu lain).")
                current_week_students = [name for name, data in st.session_state.students_data.items() if data['color'] == target_color]
                
                s_a = st.selectbox("Siswa A (Dari Minggu Ini):", ["-- Pilih Siswa A --"] + sorted(current_week_students), key="s_a")
                s_b = st.selectbox("Siswa B (Dari Minggu Lain):", ["-- Pilih Siswa B --"] + all_student_names, key="s_b")
                
                if st.button("Tukar Posisi Siswa") and s_a != "-- Pilih Siswa A --" and s_b != "-- Pilih Siswa B --":
                    color_a = st.session_state.students_data[s_a]['color']
                    color_b = st.session_state.students_data[s_b]['color']
                    st.session_state.students_data[s_a]['color'] = color_b
                    st.session_state.students_data[s_b]['color'] = color_a
                    st.success(f"Posisi {s_a} dan {s_b} berhasil ditukar!")
                    st.rerun()

        with tab_absensi:
            # Header Eksplisit agar jelas tanpa merusak tabel
            st.markdown(f"### 📋 Daftar Hadir: Minggu {selected_week}")
            st.markdown(f"<span style='color: #4A90E2; font-weight: bold; font-size: 1.1em;'>Periode: {target_date}</span>", unsafe_allow_html=True)
            
            # Tarik Data
            active_students = [
                {"Nama Siswa": name, "Kelas": data['kelas']}
                for name, data in st.session_state.students_data.items()
                if data['color'] == target_color
            ]
            active_students.sort(key=lambda x: x['Kelas'])

            if not active_students:
                st.warning("Tidak ada siswa yang dijadwalkan di minggu ini.")
            else:
                current_student_names = [s['Nama Siswa'] for s in active_students]
                
                # Reset tabel jika pindah minggu
                need_reset = False
                if st.session_state.current_viewed_week != selected_week:
                    need_reset = True
                elif not st.session_state.absensi_df.empty:
                    saved_names = st.session_state.absensi_df['Nama Siswa'].tolist()
                    if saved_names != current_student_names: need_reset = True
                else:
                    need_reset = True

                if need_reset:
                    df_data = []
                    for idx, s in enumerate(active_students):
                        df_data.append({
                            "NO": idx + 1,
                            "Nama Siswa": s['Nama Siswa'],
                            "Kelas": s['Kelas'],
                            "Senin": "-", "Selasa": "-", "Rabu": "-", "Kamis": "-", "Jumat": "-",
                            "Keterangan": ""
                        })
                    st.session_state.absensi_df = pd.DataFrame(df_data)
                    st.session_state.current_viewed_week = selected_week

                # PANEL AKSI CEPAT
                st.caption("⚡ **Aksi Cepat:** Isi tabel otomatis untuk menghemat waktu.")
                c_btn1, c_btn2, c_btn3 = st.columns(3)
                if c_btn1.button("✅ Semua Hadir (Senin-Jumat)", use_container_width=True):
                    for day in ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']: st.session_state.absensi_df[day] = '✔'
                    st.rerun()
                if c_btn2.button("📅 Set Hari Ini Saja Hadir", use_container_width=True):
                    import datetime
                    hari_ini = datetime.datetime.now().strftime("%A")
                    map_hari = {'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu', 'Thursday': 'Kamis', 'Friday': 'Jumat'}
                    if hari_ini in map_hari:
                        st.session_state.absensi_df[map_hari[hari_ini]] = '✔'
                        st.toast(f"Hadir masal diterapkan untuk hari {map_hari[hari_ini]}!")
                        st.rerun()
                    else: st.warning("Hari ini bukan hari kerja (Senin-Jumat).")
                if c_btn3.button("🧹 Bersihkan Tabel", use_container_width=True):
                    for day in ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']: st.session_state.absensi_df[day] = '-'
                    st.session_state.absensi_df['Keterangan'] = ''
                    st.rerun()

                st.markdown("<br>", unsafe_allow_html=True)
                
                # TABEL INTERAKTIF
                opsi_absen = ["✔", "S", "I", "A", "-"]
                
                edited_df = st.data_editor(
                    st.session_state.absensi_df,
                    column_config={
                        "NO": st.column_config.NumberColumn("NO", disabled=True, width="small"),
                        "Nama Siswa": st.column_config.TextColumn("Nama Siswa", disabled=True, width="medium"),
                        "Kelas": st.column_config.TextColumn("Kelas", disabled=True, width="small"),
                        "Senin": st.column_config.SelectboxColumn("Senin", options=opsi_absen, required=True),
                        "Selasa": st.column_config.SelectboxColumn("Selasa", options=opsi_absen, required=True),
                        "Rabu": st.column_config.SelectboxColumn("Rabu", options=opsi_absen, required=True),
                        "Kamis": st.column_config.SelectboxColumn("Kamis", options=opsi_absen, required=True),
                        "Jumat": st.column_config.SelectboxColumn("Jumat", options=opsi_absen, required=True),
                        "Keterangan": st.column_config.TextColumn("Keterangan Tambahan")
                    },
                    hide_index=True,
                    use_container_width=True,
                    height=min(40 * len(active_students) + 40, 500)
                )
                
                st.session_state.absensi_df = edited_df

                # DASHBOARD STATISTIK
                st.markdown("#### 📊 Statistik Kehadiran Minggu Ini")
                tot_hadir = (edited_df[['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']] == '✔').sum().sum()
                tot_sakit = (edited_df[['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']] == 'S').sum().sum()
                tot_izin = (edited_df[['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']] == 'I').sum().sum()
                tot_alfa = (edited_df[['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']] == 'A').sum().sum()
                
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Total Hadir (✔)", tot_hadir)
                m2.metric("Total Sakit (S)", tot_sakit)
                m3.metric("Total Izin (I)", tot_izin)
                m4.metric("Total Alfa (A)", tot_alfa)

                # ==========================================
                # EXPORT KE EXCEL
                # ==========================================
                st.divider()
                if st.button("💾 Generate & Download Excel (Final)", type="primary", use_container_width=True):
                    with st.spinner("Menyusun file Excel..."):
                        out_wb = openpyxl.Workbook()
                        out_ws = out_wb.active
                        out_ws.title = "ABSENSI REKAP"
                        
                        bold_font = Font(bold=True)
                        center_align = Alignment(horizontal='center', vertical='center')
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                             top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        out_ws['A1'] = "REKAP ABSENSI OJT SMK STRADA"
                        out_ws['A1'].font = Font(bold=True, size=12)
                        out_ws['A2'] = "Periode"; out_ws['C2'] = ":"; out_ws['D2'] = target_date
                        out_ws['A3'] = "Jadwal"; out_ws['C3'] = ":"; out_ws['D3'] = f"Minggu ke {selected_week}"
                        out_ws['A4'] = "Waktu"; out_ws['C4'] = ":"; out_ws['D4'] = "09:00 - 18:00"
                        out_ws['G2'] = "Hari Aktif"; out_ws['H2'] = 5
                        out_ws['G3'] = "Jumlah Siswa"; out_ws['H3'] = len(active_students)
                        
                        headers = [("NO", 1), ("Nama Siswa", 2), ("Kelas Jurusan", 3), ("Senin", 4), 
                                   ("Selasa", 5), ("Rabu", 6), ("Kamis", 7), ("Jumat", 8), 
                                   ("Kehadiran", 9), ("Jumlah Unit Dirakit", 10), ("KETERANGAN", 11)]
                        
                        for name, col_idx in headers:
                            c = out_ws.cell(row=6, column=col_idx, value=name)
                            c.font = bold_font; c.alignment = center_align; c.border = thin_border
                            
                        start_row = 7
                        for idx, row_data in edited_df.iterrows():
                            row = start_row + idx
                            
                            out_ws.cell(row=row, column=1, value=row_data['NO']).alignment = center_align
                            out_ws.cell(row=row, column=2, value=row_data['Nama Siswa'])
                            out_ws.cell(row=row, column=3, value=row_data['Kelas']).alignment = center_align
                            
                            total_hadir = 0
                            days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
                            
                            for col_offset, day in enumerate(days):
                                col_idx = 4 + col_offset
                                status = row_data[day]
                                out_ws.cell(row=row, column=col_idx, value=status).alignment = center_align
                                if status == "✔": total_hadir += 1
                            
                            out_ws.cell(row=row, column=9, value=total_hadir).alignment = center_align
                            out_ws.cell(row=row, column=10, value="")
                            
                            ket_val = str(row_data['Keterangan']) if pd.notna(row_data['Keterangan']) else ""
                            out_ws.cell(row=row, column=11, value=ket_val)
                            
                            for c_idx in range(1, 12):
                                out_ws.cell(row=row, column=c_idx).border = thin_border

                        out_ws.column_dimensions['A'].width = 5
                        out_ws.column_dimensions['B'].width = 35
                        out_ws.column_dimensions['C'].width = 15
                        for col in ['D', 'E', 'F', 'G', 'H', 'I']: 
                            out_ws.column_dimensions[col].width = 10
                        out_ws.column_dimensions['J'].width = 18
                        out_ws.column_dimensions['K'].width = 35
                        
                        buffer = io.BytesIO()
                        out_wb.save(buffer)
                        buffer.seek(0)
                        
                        st.success("✅ File Excel Siap Diunduh!")
                        st.download_button(
                            label=f"📥 Download Absensi Minggu {selected_week} (Hasil Akhir)",
                            data=buffer,
                            file_name=f"Absensi_OJT_Minggu_{selected_week}_Final.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
